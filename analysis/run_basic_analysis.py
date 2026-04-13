#!/usr/bin/env python
"""
xAquatic – Basic Exposure & Effect Analysis Script
===================================================
Server-side equivalent of PEC_GUTS_BasicAnalysis1.ipynb.

Usage
-----
python run_basic_analysis.py \\
    --mc-path   <path/to/run/exp_id/mcs/mc_id> \\
    --scenario-path <path/to/scenario> \\
    --scenario-name Muenster \\
    --output-dir <output_folder> \\
    [--run-pec true] [--run-guts true] [--exposed-only false] \\
    [--reach-ids-single 154] [--reach-ids-group 772,575,149] \\
    [--plotzoom-from 2000-05-01] [--plotzoom-to 2000-06-01]
"""

import argparse
import os
import sys
import warnings
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")  # non-interactive backend – must come before pyplot import

import matplotlib.pyplot as plt
import matplotlib.ticker as ticker

import h5py
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook

warnings.filterwarnings("ignore")


# ── logging ──────────────────────────────────────────────────────────────────
def log(severity, msg):
    """Print a structured log line (mirrors LandscapeModel log format)."""
    tag = {"ok": "OK   ", "info": "INFO ", "warn": "WARN ", "error": "ERROR"}
    print(f"{tag.get(severity, 'INFO ')} {msg}", flush=True)


# ── percentile helpers ────────────────────────────────────────────────────────
def row_percentiles(row, percentiles):
    return pd.Series(
        np.percentile(row.dropna(), [p * 100 for p in percentiles]),
        index=[f"Px{int(p * 100)}" for p in percentiles],
    )


# ── scenario-specific defaults ────────────────────────────────────────────────
SCENARIO_DEFAULTS = {
    "Rummen": {
        "reach_list_single": [1388],
        "reach_list_group": [1610, 1738, 1505],
        "plotzoom_from": "1995-05-01",
        "plotzoom_to": "1995-06-01",
    },
    "Oudebeek": {
        "reach_list_single": [42],
        "reach_list_group": [],
        "plotzoom_from": "",
        "plotzoom_to": "",
    },
    "Oudebeek-Beek7": {
        "reach_list_single": [42],
        "reach_list_group": [],
        "plotzoom_from": "",
        "plotzoom_to": "",
    },
    "Muenster": {
        "reach_list_single": [154],
        "reach_list_group": [772, 575, 149],
        "plotzoom_from": "2000-05-01",
        "plotzoom_to": "2000-06-01",
    },
        "Wetter_2": {
            "reach_list_single": [154],
            "reach_list_group": [131, 130, 437],
            "plotzoom_from": "1991-05-01",
            "plotzoom_to": "1991-05-10",
        },
    "Muschenheim": {
        "reach_list_single": [315],
        "reach_list_group": [151, 149, 735],
        "plotzoom_from": "",
        "plotzoom_to": "",
    },
    "Bruchenbruecken": {
        "reach_list_single": [660],
        "reach_list_group": [652, 653, 4443],
        "plotzoom_from": "1991-05-01",
        "plotzoom_to": "1991-06-01",
    },
    "Funne": {
        "reach_list_single": [12],
        "reach_list_group": [99, 97, 167],
        "plotzoom_from": "2010-05-01",
        "plotzoom_to": "2010-06-01",
    },
    "GKB": {
        "reach_list_single": [166],
        "reach_list_group": [12, 144, 165],
        "plotzoom_from": "2010-04-01",
        "plotzoom_to": "2010-07-01",
    },
}

# ── exposure model configuration ─────────────────────────────────────────────
EXPOSURE_MODEL_CONFIG = {
    "CascadeToxswa": {
        "pecsw_key":    "CascadeToxswa/ConLiqWatTgtAvg",
        "pecsed_key":   "CascadeToxswa/CntSedTgt1",
        "pecsw_scale":  1_000_000,   # mg/L → ng/L
        "pecsed_scale": 1_000,       # mg/kg → µg/kg
        "effect_prefix": "CascadeToxswa",
    },
    "StepsRiverNetwork": {
        "pecsw_key":    "StepsRiverNetwork/PEC_SW",
        "pecsed_key":   "StepsRiverNetwork/PEC_SED",
        "pecsw_scale":  1_000,       # µg/L → ng/L
        "pecsed_scale": 1_000,       # mg/kg → µg/kg
        "effect_prefix": "StepsRiverNetwork",
    },
}
VALID_EXPOSURE_MODELS = tuple(EXPOSURE_MODEL_CONFIG.keys())


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="xAquatic Basic Exposure & Effect Analysis")
    ap.add_argument("--mc-path", required=True,
                    help="Path to MC run folder (…/mcs/<mc_id>)")
    ap.add_argument("--scenario-path", required=True,
                    help="Path to scenario folder")
    ap.add_argument("--scenario-name", default="",
                    help="Scenario name for built-in reach-list defaults")
    ap.add_argument("--output-dir", required=True,
                    help="Folder where results (Excel + PNG) are written")
    ap.add_argument("--run-pec",
                    type=lambda x: x.lower() not in ("false", "0", "no"),
                    default=True)
    ap.add_argument("--run-guts",
                    type=lambda x: x.lower() not in ("false", "0", "no"),
                    default=True)
    ap.add_argument("--exposed-only",
                    type=lambda x: x.lower() in ("true", "1", "yes"),
                    default=False)
    ap.add_argument("--reach-ids-single", default="",
                    help="Comma-separated reach IDs for single-reach time-series")
    ap.add_argument("--reach-ids-group", default="",
                    help="Comma-separated reach IDs for grouped time-series")
    ap.add_argument("--plotzoom-from", default="",
                    help="Time-series zoom start date (YYYY-MM-DD)")
    ap.add_argument("--plotzoom-to", default="",
                    help="Time-series zoom end date (YYYY-MM-DD)")
    ap.add_argument("--exposure-model", default="CascadeToxswa",
                    choices=VALID_EXPOSURE_MODELS,
                    help="Exposure model whose HDF5 outputs are analysed "
                         f"({', '.join(VALID_EXPOSURE_MODELS)})")
    args = ap.parse_args()

    # ── path setup ────────────────────────────────────────────────────────────
    mc_path       = Path(args.mc_path)
    scenario_path = Path(args.scenario_path)
    output_dir    = Path(args.output_dir)
    scenario_name = args.scenario_name

    # ── exposure model config ─────────────────────────────────────────────────
    exposure_model = args.exposure_model
    cfg = EXPOSURE_MODEL_CONFIG[exposure_model]

    defaults = SCENARIO_DEFAULTS.get(scenario_name, {})

    def _parse_ids(raw, default):
        ids = [int(x.strip()) for x in raw.split(",") if x.strip()]
        return ids if ids else default

    reach_list_single = _parse_ids(args.reach_ids_single,
                                   defaults.get("reach_list_single", []))
    reach_list_group  = _parse_ids(args.reach_ids_group,
                                   defaults.get("reach_list_group", []))
    plotzoom_from = args.plotzoom_from or defaults.get("plotzoom_from", "")
    plotzoom_to   = args.plotzoom_to   or defaults.get("plotzoom_to", "")

    # ── extract experiment / MC IDs from path ─────────────────────────────────
    parts   = list(mc_path.parts)
    mc_run  = mc_path.name
    exp_id  = "Experiment"
    try:
        mcs_idx = parts.index("mcs")
        mc_run  = parts[mcs_idx + 1]
        exp_id  = parts[mcs_idx - 1]
    except (ValueError, IndexError):
        pass

    h5_data = mc_path / "store" / "arr.dat"

    # ── analysis settings ─────────────────────────────────────────────────────
    pec_temporal_percentiles  = [0.1, 0.25, 0.5, 0.75, 0.9, 0.95, 0.99, 1.0]
    pec_spatial_percentiles   = [0.1, 0.25, 0.5, 0.75, 0.9, 0.95, 0.99, 1.0]
    guts_temporal_percentiles = [0.01, 0.05, 0.1, 0.25, 0.5, 0.75, 0.9, 0.95, 0.99, 1.0]
    guts_spatial_percentiles  = [0.01, 0.05, 0.1, 0.25, 0.5, 0.75, 0.9, 0.95, 0.99, 1.0]
    boxplot_temporal_percentile = [0.9, 0.99, 1.0]
    lower_LP50_threshold, upper_LP50_threshold = 0.01, 1000

    # ── prepare output folder ─────────────────────────────────────────────────
    output_dir.mkdir(parents=True, exist_ok=True)
    os.chdir(output_dir)

    timestamp_str  = datetime.now().strftime("%Y%m%d%H%M%S")
    excel_filename = f"{exp_id}_{mc_run[:8]}_{timestamp_str}.xlsx"
    pd.DataFrame().to_excel(excel_filename, index=False)

    log("info",  f"Experiment  : {exp_id}")
    log("info",  f"MC Run      : {mc_run}")
    log("info",  f"Scenario    : {scenario_name or '(none)'}")
    log("info",  f"HDF5 store  : {h5_data}")
    log("info",  f"Output dir  : {output_dir}")
    log("info",  f"Excel file  : {excel_filename}")
    log("info",  f"Run PEC     : {args.run_pec}")
    log("info",  f"Run GUTS    : {args.run_guts}")
    log("info",  f"Exposure model: {exposure_model}")

    if not h5_data.exists():
        log("error", f"HDF5 store not found: {h5_data}")
        sys.exit(1)

    # ── load geo / scenario shapefile (optional) ──────────────────────────────
    df_geo_reaches            = None
    df_scenario_geo_attributes = None
    shp_path = scenario_path / "geo" / "Reachlist_shp.shp"
    try:
        import geopandas as gpd
        if shp_path.exists():
            df_geo_reaches             = gpd.read_file(str(shp_path))
            df_scenario_geo_attributes = (
                df_geo_reaches.drop(columns="geometry")
                              .set_index("key")
            )
            log("ok",   f"Shapefile loaded: {shp_path.name}")
        else:
            log("info", f"Shapefile not found – geo plots skipped ({shp_path})")
    except ImportError:
        log("info", "geopandas not available – geo plots skipped")

    # ─────────────────────────────────────────────────────────────────────────
    # EXPOSURE ANALYSIS
    # ─────────────────────────────────────────────────────────────────────────
    df_pecsw = df_pecsed = None
    pecsw_perc_by_reach = pecsed_perc_by_reach = None

    if args.run_pec:
        log("info", "=== Exposure Analysis ===")
        try:
            with h5py.File(h5_data, "r") as f:
                if cfg["pecsw_key"] not in f:
                    raise KeyError(f"{cfg['pecsw_key']} not found in HDF5")
                reach_ids_pecsw = f[
                    f[cfg["pecsw_key"]].attrs["dim1_element_names"]
                ][:]
                starttime_pecsw = f[cfg["pecsw_key"]].attrs["dim0_offset"]
                df_pecsw  = pd.DataFrame(f[cfg["pecsw_key"]][:])  * cfg["pecsw_scale"]
                df_pecsed = pd.DataFrame(f[cfg["pecsed_key"]][:]) * cfg["pecsed_scale"]

            ti = pd.date_range(starttime_pecsw, periods=len(df_pecsw), freq="h")
            for df in (df_pecsw, df_pecsed):
                df.insert(0, "time", ti)
                df.set_index("time", inplace=True)
            df_pecsw.columns  = reach_ids_pecsw
            df_pecsed.columns = reach_ids_pecsw
            log("ok", f"PECsw loaded: {df_pecsw.shape[0]} timesteps × {df_pecsw.shape[1]} reaches")

            # filter to exposed streams
            if args.exposed_only:
                for df_ref, df_obj in [("pecsw", df_pecsw), ("pecsed", df_pecsed)]:
                    valid_segs  = df_obj.gt(0).any(axis=0)
                    valid_times = df_obj.gt(0).any(axis=1)
                    if df_ref == "pecsw":
                        df_pecsw  = df_pecsw.loc[valid_times, valid_segs]
                    else:
                        df_pecsed = df_pecsed.loc[valid_times, valid_segs]
                log("info", f"Exposed-only: {df_pecsw.shape[1]} stream segments retained")

            # ── percentile tables ─────────────────────────────────────────────
            log("info", "Computing PEC percentile tables…")
            pecsw_perc_by_reach  = df_pecsw.quantile(pec_temporal_percentiles)
            pecsed_perc_by_reach = df_pecsed.quantile(pec_temporal_percentiles)
            PECsw_x_t  = pecsw_perc_by_reach.apply(
                lambda r: row_percentiles(r, pec_spatial_percentiles), axis=1)
            PECsed_x_t = pecsed_perc_by_reach.apply(
                lambda r: row_percentiles(r, pec_spatial_percentiles), axis=1)

            num_fmt_pec = "0.0000"
            with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                PECsw_x_t.to_excel(writer,  sheet_name="PECsw_percentiles_x_t")
                PECsed_x_t.to_excel(writer, sheet_name="PECsed_percentiles_x_t")
                for sn in ("PECsw_percentiles_x_t", "PECsed_percentiles_x_t"):
                    ws = writer.sheets[sn]
                    ws.cell(row=1, column=1).value = "perc_x: right / perc_t: down"
                    for row in ws.iter_rows(min_row=2, min_col=2, max_col=30, max_row=30):
                        for cell in row:
                            cell.number_format = num_fmt_pec
            log("ok", "PEC percentile tables exported to Excel")

            # write PECsw table as JSON for web display
            import json as _json
            pecsw_json = {
                "columns": list(PECsw_x_t.columns),
                "rows": [
                    {"pt": round(float(idx), 4),
                     "values": [round(float(v), 4) for v in row]}
                    for idx, row in PECsw_x_t.iterrows()
                ],
                "unit": "ng/L",
            }
            with open("pecsw_table.json", "w") as _jf:
                _json.dump(pecsw_json, _jf)
            log("ok", "PECsw table written to pecsw_table.json")

            # ── single-reach time series ──────────────────────────────────────
            if reach_list_single:
                rid = reach_list_single[0]
                if rid in df_pecsw.columns:
                    fig, ax = plt.subplots(figsize=(12, 5))
                    ax.plot(df_pecsw.index, df_pecsw[rid], linewidth=0.7)
                    ax.set_title(f"PECsw Time Series – Reach {rid}")
                    ax.set_xlabel("Time"); ax.set_ylabel("PECsw [ng/L]")
                    ax.grid(True); fig.tight_layout()
                    fname = f"PECsw_timeSeries_full_r{rid}.png"
                    fig.savefig(fname, dpi=150); plt.close(fig)
                    log("ok", f"Saved: {fname}")

                    # zoomed
                    if plotzoom_from and plotzoom_to:
                        subset = df_pecsw[plotzoom_from:plotzoom_to]
                        if len(subset):
                            fig2, ax2 = plt.subplots(figsize=(12, 5))
                            ax2.plot(subset.index, subset[rid], linewidth=1.0)
                            ax2.set_title(f"PECsw Time Series (zoomed) – Reach {rid}")
                            ax2.set_xlabel("Time"); ax2.set_ylabel("PECsw [ng/L]")
                            ax2.grid(True); fig2.tight_layout()
                            fname2 = f"PECsw_timeSeries_zoomed_r{rid}.png"
                            fig2.savefig(fname2, dpi=150); plt.close(fig2)
                            log("ok", f"Saved: {fname2}")

                    if rid in df_pecsed.columns:
                        fig3, ax3 = plt.subplots(figsize=(12, 5))
                        ax3.plot(df_pecsed.index, df_pecsed[rid],
                                 linewidth=0.7, color="sienna")
                        ax3.set_title(f"PECsed Time Series – Reach {rid}")
                        ax3.set_xlabel("Time"); ax3.set_ylabel("PECsed [µg/kg]")
                        ax3.grid(True); fig3.tight_layout()
                        fname3 = f"PECsed_timeSeries_full_r{rid}.png"
                        fig3.savefig(fname3, dpi=150); plt.close(fig3)
                        log("ok", f"Saved: {fname3}")
                else:
                    log("warn", f"Reach {rid} not found in PECsw data")

            # ── group time series ─────────────────────────────────────────────
            if reach_list_group:
                valid_group = [r for r in reach_list_group if r in df_pecsw.columns]
                if valid_group:
                    fig, ax = plt.subplots(figsize=(12, 5))
                    for r in valid_group:
                        ax.plot(df_pecsw.index, df_pecsw[r], linewidth=0.7, label=str(r))
                    ax.legend(title="Reach-ID", fontsize="small")
                    ax.set_title("PECsw Time Series – Multiple Reaches")
                    ax.set_xlabel("Time"); ax.set_ylabel("PECsw [ng/L]")
                    ax.grid(True); fig.tight_layout()
                    fname = "PECsw_timeSeries_group.png"
                    fig.savefig(fname, dpi=150); plt.close(fig)
                    log("ok", f"Saved: {fname}")

            # ── Strahler boxplots ─────────────────────────────────────────────
            if (df_scenario_geo_attributes is not None
                    and "strahler" in df_scenario_geo_attributes.columns):
                pecsw_T = pecsw_perc_by_reach.transpose()
                pecsw_T.index.name = "ReachID"
                pecsw_T = pecsw_T.join(df_scenario_geo_attributes["strahler"])
                sns.set_theme(style="whitegrid")
                for bp in boxplot_temporal_percentile:
                    if bp in pecsw_T.columns:
                        fig, ax = plt.subplots(figsize=(10, 5))
                        sns.boxplot(x="strahler", y=bp, hue="strahler",
                                    data=pecsw_T, palette="viridis",
                                    legend=False, ax=ax)
                        pct = int(bp * 100)
                        ax.set_title(
                            f"PECsw {pct}th Percentile over Time – by Strahler Order")
                        ax.set_xlabel("Strahler Order"); ax.set_ylabel("PECsw [ng/L]")
                        fig.tight_layout()
                        fname = f"PECsw_{pct}th_by_Strahler.png"
                        fig.savefig(fname, dpi=150); plt.close(fig)
                        log("ok", f"Saved: {fname}")

            # ── map (no basemap – offline safe) ───────────────────────────────
            if df_geo_reaches is not None and pecsw_perc_by_reach is not None:
                try:
                    pecsw_T = pecsw_perc_by_reach.transpose()
                    pecsw_T.index.name = "ReachID"
                    col_99 = 0.99 if 0.99 in pecsw_T.columns else pecsw_T.columns[-1]
                    gdf = df_geo_reaches.merge(
                        pecsw_T[[col_99]], left_on="key", right_on="ReachID", how="left")
                    fig, ax = plt.subplots(figsize=(10, 8))
                    gdf.plot(column=col_99, cmap="viridis", legend=True, ax=ax,
                             missing_kwds={"color": "lightgrey"})
                    ax.set_title(f"PECsw – {int(col_99 * 100)}th Percentile over Time")
                    ax.axis("off"); fig.tight_layout()
                    fname = "PECsw_Px99_Map.png"
                    fig.savefig(fname, dpi=150); plt.close(fig)
                    log("ok", f"Saved: {fname}")
                except Exception as exc:
                    log("warn", f"PECsw map failed: {exc}")

        except Exception as exc:
            log("error", f"Exposure analysis failed: {exc}")
            import traceback; traceback.print_exc(file=sys.stdout)

    # ─────────────────────────────────────────────────────────────────────────
    # EFFECT ANALYSIS
    # ─────────────────────────────────────────────────────────────────────────
    if args.run_guts:
        log("info", "=== Effect Analysis ===")
        try:
            # Map of (internal_key, hdf5_path, is_survival_cube)
            _ep = cfg["effect_prefix"]
            GUTS_KEYS = [
                ("surv_sd_sp1", f"IndEffect_{_ep}_SD_Species1/GutsSurvivalReaches", True),
                ("surv_sd_sp2", f"IndEffect_{_ep}_SD_Species2/GutsSurvivalReaches", True),
                ("surv_sd_sp3", f"IndEffect_{_ep}_SD_Species3/GutsSurvivalReaches", True),
                ("surv_it_sp1", f"IndEffect_{_ep}_IT_Species1/GutsSurvivalReaches", True),
                ("surv_it_sp2", f"IndEffect_{_ep}_IT_Species2/GutsSurvivalReaches", True),
                ("surv_it_sp3", f"IndEffect_{_ep}_IT_Species3/GutsSurvivalReaches", True),
                ("lp50_sd_sp1", f"IndEffect_LP50_{_ep}_SD_Species1/LP50", False),
                ("lp50_sd_sp2", f"IndEffect_LP50_{_ep}_SD_Species2/LP50", False),
                ("lp50_sd_sp3", f"IndEffect_LP50_{_ep}_SD_Species3/LP50", False),
                ("lp50_it_sp1", f"IndEffect_LP50_{_ep}_IT_Species1/LP50", False),
                ("lp50_it_sp2", f"IndEffect_LP50_{_ep}_IT_Species2/LP50", False),
                ("lp50_it_sp3", f"IndEffect_LP50_{_ep}_IT_Species3/LP50", False),
            ]

            dfs = {}
            with h5py.File(h5_data, "r") as f:
                # get spatial reference from SD Species 1 survival
                ref_surv_key = f"IndEffect_{_ep}_SD_Species1/GutsSurvivalReaches"
                ref_lp50_key = f"IndEffect_LP50_{_ep}_SD_Species1/LP50"
                if ref_surv_key not in f:
                    raise KeyError(f"{ref_surv_key} not in HDF5 – GUTS data not present")

                starttime_guts = f[ref_surv_key].attrs["dim0_offset"]
                reach_ids_guts = f[f[ref_surv_key].attrs["dim1_element_names"]][:]
                if ref_lp50_key in f:
                    reach_ids_lp50 = f[f[ref_lp50_key].attrs["dim1_element_names"]][:]
                else:
                    reach_ids_lp50 = reach_ids_guts

                for key, hdf_path, is_surv in GUTS_KEYS:
                    if hdf_path in f:
                        if is_surv:
                            dfs[key] = pd.DataFrame(f[hdf_path][:, :, 10])
                        else:
                            dfs[key] = pd.DataFrame(f[hdf_path][:])

            log("ok", f"GUTS datasets loaded: {len(dfs)}")

            # attach time index and reach IDs
            n_years_surv = len(dfs.get("surv_sd_sp1", pd.DataFrame()))
            n_years_lp50 = len(dfs.get("lp50_sd_sp1", pd.DataFrame()))
            ti_surv = pd.date_range(datetime(starttime_guts, 1, 1),
                                    periods=n_years_surv, freq="YE")
            ti_lp50 = pd.date_range(datetime(starttime_guts, 1, 1),
                                    periods=n_years_lp50, freq="YE")

            for key, _, is_surv in GUTS_KEYS:
                if key not in dfs:
                    continue
                ti   = ti_surv if is_surv else ti_lp50
                rids = reach_ids_guts if is_surv else reach_ids_lp50
                n    = len(dfs[key])
                dfs[key].insert(0, "time", ti[:n])
                dfs[key].set_index("time", inplace=True)
                dfs[key].columns = rids

            # fix LP50 values
            for key in [k for k in dfs if k.startswith("lp50")]:
                dfs[key][dfs[key] < 0] = np.nan
                dfs[key][dfs[key] < lower_LP50_threshold] = lower_LP50_threshold
                dfs[key][dfs[key] > upper_LP50_threshold] = upper_LP50_threshold

            # filter to exposed streams
            if args.exposed_only and df_pecsw is not None:
                exposed = df_pecsw.columns[(df_pecsw > 0).any(axis=0)]
                for key in dfs:
                    common = [r for r in exposed if r in dfs[key].columns]
                    dfs[key] = dfs[key][common]
                log("info", f"GUTS filtered to {len(exposed)} exposed reaches")

            # ── percentile tables ─────────────────────────────────────────────
            log("info", "Computing GUTS percentile tables…")
            GUTS_SHEET_MAP = {
                "surv_sd_sp1": "GUTSperc_SurvSdSp1",
                "surv_sd_sp2": "GUTSperc_SurvSdSp2",
                "surv_sd_sp3": "GUTSperc_SurvSdSp3",
                "surv_it_sp1": "GUTSperc_SurvItSp1",
                "surv_it_sp2": "GUTSperc_SurvItSp2",
                "surv_it_sp3": "GUTSperc_SurvItSp3",
                "lp50_sd_sp1": "GUTSperc_LP50SdSp1",
                "lp50_sd_sp2": "GUTSperc_LP50SdSp2",
                "lp50_sd_sp3": "GUTSperc_LP50SdSp3",
                "lp50_it_sp1": "GUTSperc_LP50ItSp1",
                "lp50_it_sp2": "GUTSperc_LP50ItSp2",
                "lp50_it_sp3": "GUTSperc_LP50ItSp3",
            }
            num_fmt_guts = "0.000"
            with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                for key, sheet in GUTS_SHEET_MAP.items():
                    if key not in dfs:
                        continue
                    perc_by_reach = dfs[key].quantile(guts_temporal_percentiles)
                    perc_x_t = perc_by_reach.apply(
                        lambda r: row_percentiles(r, guts_spatial_percentiles), axis=1)
                    perc_x_t.to_excel(writer, sheet_name=sheet)
                    ws = writer.sheets[sheet]
                    ws.cell(row=1, column=1).value = "perc_x: right / perc_t: down"
                    for row in ws.iter_rows(min_row=2, min_col=2, max_col=30, max_row=30):
                        for cell in row:
                            cell.number_format = num_fmt_guts
            log("ok", "GUTS percentile tables exported to Excel")

            # ── LP50 histogram & Strahler heatmap ─────────────────────────────
            if "lp50_sd_sp1" in dfs:
                try:
                    lp50_T = dfs["lp50_sd_sp1"].transpose()
                    lp50_T.index.name = "ReachID"
                    lp50_T.columns = [
                        str(c.year) if isinstance(c, pd.Timestamp) else str(c)
                        for c in lp50_T.columns
                    ]

                    if (df_scenario_geo_attributes is not None
                            and "strahler" in df_scenario_geo_attributes.columns):
                        lp50_T = lp50_T.join(df_scenario_geo_attributes["strahler"], how="left")

                    first_year_col = lp50_T.columns[0]
                    sns.set_theme(style="whitegrid")
                    fig, ax = plt.subplots(figsize=(12, 5))
                    if "strahler" in lp50_T.columns:
                        sns.histplot(data=lp50_T, x=first_year_col,
                                     hue="strahler", multiple="stack", bins=50, ax=ax)
                    else:
                        sns.histplot(data=lp50_T, x=first_year_col, bins=50, ax=ax)
                    ax.set_title(f"LP50 Distribution – SD Species 1 (year {first_year_col})")
                    ax.xaxis.set_major_locator(ticker.MultipleLocator(100))
                    fig.tight_layout()
                    fname = "LP50_Histogram_SD_Species1.png"
                    fig.savefig(fname, dpi=150); plt.close(fig)
                    log("ok", f"Saved: {fname}")

                    # Strahler heatmap
                    if "strahler" in lp50_T.columns:
                        year_cols = [c for c in lp50_T.columns if c != "strahler"]
                        bins = [-float("inf"), 0.1, 1, 10, 100, 200, 500, 1000]
                        bin_labels = ["<0.1", "[0.1,1)", "[1,10)", "[10,100)",
                                      "[100,200)", "[200,500)", "[500,1000]"]
                        grouped = {}
                        for col in year_cols:
                            cnts = lp50_T.groupby("strahler")[col].apply(
                                lambda x: pd.cut(x, bins=bins, labels=bin_labels,
                                                 include_lowest=True).value_counts()
                            ).unstack().fillna(0)
                            grouped[col] = cnts
                        combined = pd.concat(grouped, axis=1)
                        features = list(combined.columns.get_level_values(0).unique())
                        nc = min(3, len(features))
                        nr = int(np.ceil(len(features) / nc))
                        fig2, axes = plt.subplots(nr, nc,
                                                   figsize=(nc * 5, nr * 4),
                                                   squeeze=False)
                        axes_flat = axes.flatten()
                        for i, feat in enumerate(features):
                            if i < len(axes_flat):
                                sns.heatmap(combined[feat], annot=True, fmt="d",
                                            cmap="YlGnBu", ax=axes_flat[i], cbar=False)
                                axes_flat[i].set_title(f"Year {feat}")
                                axes_flat[i].set_xlabel("LP50 Bin")
                                axes_flat[i].set_ylabel("Strahler Order")
                        for j in range(len(features), len(axes_flat)):
                            fig2.delaxes(axes_flat[j])
                        fig2.suptitle(f"LP50 by Strahler – SD Species 1 – {exp_id}",
                                      fontsize=12)
                        fig2.tight_layout()
                        fname2 = "LP50_Strahler_Heatmap_SD_Species1.png"
                        fig2.savefig(fname2, dpi=150); plt.close(fig2)
                        log("ok", f"Saved: {fname2}")
                except Exception as exc:
                    log("warn", f"LP50 plots failed: {exc}")

            # ── GUTS survival map (last simulation year) ──────────────────────
            if df_geo_reaches is not None and "surv_sd_sp1" in dfs:
                try:
                    surv_T = dfs["surv_sd_sp1"].transpose()
                    surv_T.index.name = "ReachID"
                    surv_T.columns = [
                        str(c.year) if isinstance(c, pd.Timestamp) else str(c)
                        for c in surv_T.columns
                    ]
                    last_yr = surv_T.columns[-1]
                    gdf = df_geo_reaches.merge(
                        surv_T[[last_yr]], left_on="key", right_on="ReachID", how="left")
                    fig, ax = plt.subplots(figsize=(10, 8))
                    gdf.plot(column=last_yr, cmap="RdYlGn", legend=True, ax=ax,
                             missing_kwds={"color": "lightgrey"}, vmin=0, vmax=1)
                    ax.set_title(f"GUTS Survival (SD Species 1) – Year {last_yr}")
                    ax.axis("off"); fig.tight_layout()
                    fname = f"GUTS_Survival_SD_Sp1_{last_yr}_Map.png"
                    fig.savefig(fname, dpi=150); plt.close(fig)
                    log("ok", f"Saved: {fname}")
                except Exception as exc:
                    log("warn", f"GUTS survival map failed: {exc}")

        except Exception as exc:
            log("error", f"Effect analysis failed: {exc}")
            import traceback; traceback.print_exc(file=sys.stdout)

    log("ok", f"Analysis complete. Results in: {output_dir}")
    log("info", f"Excel file:  {excel_filename}")


if __name__ == "__main__":
    main()
